from openpyxl import load_workbook
from .base import BaseConverter

class ExcelConverter(BaseConverter):

    def convert(self, path: str) -> str:
        wb = load_workbook(path, data_only=True)
        md_blocks = []

        for sheet in wb.sheetnames:
            ws = wb[sheet]

            md_blocks.append(f"# Лист: {sheet}")
            md_blocks.append(self._sheet_to_md(ws))
            md_blocks.append("")  # пустая строка между листами

        return "\n".join(md_blocks)

    def _sheet_to_md(self, ws):
        rows = list(ws.rows)

        if not rows:
            return "_Пустой лист_"

        # Преобразуем строки в списки значений
        table = []
        for row in rows:
            table.append([self._cell_to_str(cell) for cell in row])

        # Если нет заголовка — создаём пустой
        headers = table[0]
        body = table[1:]

        md = []
        # Header row
        md.append("| " + " | ".join(headers) + " |")
        md.append("| " + " | ".join("---" for _ in headers) + " |")

        # Body
        for row in body:
            md.append("| " + " | ".join(row) + " |")

        return "\n".join(md)

    def _cell_to_str(self, cell):
        v = cell.value
        if v is None:
            return ""
        return str(v).replace("\n", " ").strip()
